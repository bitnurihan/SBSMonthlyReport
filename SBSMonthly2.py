from datetime import datetime, date
from openpyxl import load_workbook
from copy import copy
import xlrd


def read_data_from_excel(write_file_name):
    global read_excel_file, worksheet_read, current_month, row, col, array, inner_array, i, j, value
    read_excel_file = xlrd.open_workbook(write_file_name)
    worksheet_read = read_excel_file.sheet_by_name('전체 수도권 (P) - 가구 (1408)')
    row = 3
    col = 4
    array = []
    inner_array = []
    for i in range(1):
        for j in range(10):
            value = worksheet_read.cell_value(rowx=row, colx=col)
            inner_array.append(value)
            col += 1

        array.append(inner_array)
        inner_array = []
        col = 3
        row += 1
        print()


def paste_to_excel(row, col):
    global i, j, test, test
    for i in range(1):
        for j in range(10):
            test = worksheet_write.cell(row=row, column=col)
            test.value = array[i][j]
            col += 1
        print()


def get_work_line(work_sheet_name):

    if work_sheet_name == r'기존전시간대시청률(06-11,17-24)':
        start_line = 376
    elif work_sheet_name == r'추가전시간대시청률(06-25)':
        start_line = 116
    else:
        return

    init_year = 2019
    current_year = datetime.now().year
    current_month = datetime.now().month
    extra_line = 0
    if current_month > 3:
        extra_line = 2
    elif current_month > 6:
        extra_line = 4
    elif current_month > 9:
        extra_line = 6

    return start_line + (current_year - init_year) * 34 + (current_month - 1) * 2 + extra_line


def get_work_line_2():  # 자사케이블 시청률을 위한 work_line setting 툴
    start_line = 170
    init_year = 2019
    current_year = datetime.now().year
    current_month = datetime.now().month
    if current_month < 4:
        current_month = 2
    elif current_month < 7:
        current_month = 3
    elif current_month < 10:
        current_month = 4
    elif current_month <= 12:
        current_month = 5

    return start_line + (current_year - init_year) * 12 + (current_month - 1) * 2


def hidden_cells():
    global col
    for col in ['N', 'O', 'P']:
        worksheet_write.column_dimensions[col].hidden = True


def get_zero_month():
    month = (datetime.now().month - 1)

    if month < 10:
        month = "0" + str(month)

    return str(month)


def write_formulas():
    global row, col
    row = get_work_line(work_sheet_name=worksheet_write.title)
    col = 1
    alphabet_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']
    for colunm in range(2, 13):
        formulas = '=%s%d/$N$%d' % (alphabet_list[colunm], row, row)
        formulas_cell = worksheet_write.cell(row + 1, col + 2)
        formulas_cell.value = formulas
        col += 1

    worksheet_write.cell(row, 13).value = '=N%d-SUM(C%d:L%d)+J%d' % (row, row, row, row)

    worksheet_write.cell(row, 15).value = '=SUM(C%d:M%d)-J%d' % (row, row, row)

    worksheet_write.cell(row, 16).value = '=O%d=N%d' % (row, row)

    worksheet_write.cell(row+1, 15).value = '=SUM(C%d:M%d)-J%d' % (row+1, row+1, row+1)


def get_style():
    global row, col, i, j
    row = get_work_line(work_sheet_name=worksheet_write.title)
    col = 1
    for i in range(2):
        for j in range(16):
            worksheet_write.cell(row, col)._style = copy(worksheet_write.cell(row - 2, col)._style)
            col += 1

        row += 1
        col = 1


def get_style_2():  # 자사케이블 시청률을 위한 style setting 툴/ 2월에만 작업 됨
    global row, col, i, j

    copy_formulas(8)
    copy_formulas(9)

    row = get_work_line_2()  # 선 스타일 복사
    col = 1
    for i in range(10):
        for j in range(6):
            worksheet_write.cell(row, col)._style = copy(worksheet_write.cell(row - 10, col)._style)
            col += 1

        row += 1
        col = 1

    row = get_work_line_2()  # Labeling 같이 복사 해오기
    for i in range(8):
        for j in range(2):
            worksheet_write.cell(row, col+1).value = copy(worksheet_write.cell(row - 10, col+1).value)
            col += 1

        row += 1
        col = 1

    (worksheet_write.cell(row=get_work_line_2(), column=1)).value = datetime.now().year
    (worksheet_write.cell(row=get_work_line_2()+8, column=1)).value = "%d년 연간" % datetime.now().year
    (worksheet_write.cell(row=get_work_line_2()+8, column=3)).value = "Viewership"
    (worksheet_write.cell(row=get_work_line_2()+9, column=3)).value = "Market Share"


def copy_formulas(x1):
    global row, col
    row = get_work_line_2() + x1
    col = 4
    alphabet_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                     'U', 'V', 'W', 'X', 'Y', 'Z']
    for colunm in range(3, 6):
        formulas = '=AVERAGE(%s%d,%s%d,%s%d,%s%d)' % (alphabet_list[colunm], row - 8, alphabet_list[colunm], row - 6,
                                                      alphabet_list[colunm], row - 4, alphabet_list[colunm], row - 2)
        formulas_cell = worksheet_write.cell(row, col)
        formulas_cell.value = formulas
        col += 1


def setting_date():
    date_cell = worksheet_write.cell(row=(get_work_line(work_sheet_name=worksheet_write.title)), column=1)
    date_cell.value = "{}.{}".format(datetime.now().year, get_zero_month())


def cell_labeling():
    (worksheet_write.cell(row=get_work_line(work_sheet_name=worksheet_write.title), column=2)).value = "Viewership"  # cell마다 viewership 넣기
    (worksheet_write.cell(row=get_work_line(work_sheet_name=worksheet_write.title) + 1, column=2)).value = "Market Share"  # cell마다 Market Share 넣기


###1. 기존전시간대시청률(06-11,17-24)
write_excel_file = load_workbook(filename=r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\MonthlyReport2.xlsx')

read_data_from_excel(r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\1.xls')
worksheet_write = write_excel_file[r'기존전시간대시청률(06-11,17-24)']

paste_to_excel(get_work_line(work_sheet_name=worksheet_write.title), 3)

get_style()  # 셀 스타일 복사해서 붙여넣기

row = get_work_line(work_sheet_name=worksheet_write.title)
worksheet_write.cell(row, 14).value = copy(worksheet_read.cell(rowx=3, colx=14).value)  # N값 넣기

write_formulas()  # 함수넣기

hidden_cells()  # 셀숨기기

cell_labeling()  # cell마다 viewership, Market Share 넣기

setting_date()  # 날짜넣기


###2. 추가전시간대시청률(06-25)

read_data_from_excel(r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\1_1.xls')
worksheet_write = write_excel_file[r'추가전시간대시청률(06-25)']

paste_to_excel(get_work_line(work_sheet_name=worksheet_write.title), 3)

get_style()  # 셀 스타일 복사해서 붙여넣기

row = get_work_line(work_sheet_name=worksheet_write.title)
worksheet_write.cell(row, 14).value = copy(worksheet_read.cell(rowx=3, colx=14).value)  # N값 넣기

write_formulas()  # 함수넣기

hidden_cells()  # 셀숨기기

cell_labeling()  # cell마다 viewership, Market Share 넣기

setting_date()  # 날짜넣기


###3. 자사케이블 시청률
read_excel_file = xlrd.open_workbook(r'C:\Users\hanbi01\Desktop\한빛누리\(매월)SBS월간업데이트\1_2.xls')
worksheet_read = read_excel_file.sheet_by_name('전체 수도권 (P) - CATV가구(N) (')
worksheet_write = write_excel_file[r'자사케이블 시청률']

if datetime.now().month == 2:
    get_style_2()  # 스타일 복사하기

row = 3
col = 4
array = []
inner_array = []
for i in range(2):
    for j in range(3):
        value = worksheet_read.cell_value(rowx=row, colx=col)
        inner_array.append(value)
        col += 1

    array.append(inner_array)
    inner_array = []
    col = 4
    row += 1
    print(array)

row = get_work_line_2()
col = 4

for i in range(2):
    for j in range(3):
        test = worksheet_write.cell(row=row, column=col)
        test.value = array[i][j]
        col += 1
    row += 1
    col = 4
    print()

write_excel_file.save('testfile2.xlsx')
